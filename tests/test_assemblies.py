import pytest

from digisearch.web.core import FeatureRegistry
from digisearch.web.core.db import Database
from digisearch.web.features.assemblies import feature as assemblies_feature
from digisearch.web.features.assemblies import importer, repo
from digisearch.web.features.catalog import feature as catalog_feature


@pytest.fixture
def db(tmp_path):
    database = Database(tmp_path / "a.db")
    reg = FeatureRegistry()
    reg.register(catalog_feature, assemblies_feature)
    database.apply_migrations(reg)
    return database


def _part(db, part_no, kind, minimrp_id):
    with db.connect() as conn:
        return conn.execute(
            "INSERT INTO parts (part_no, kind, minimrp_id) VALUES (?, ?, ?)",
            (part_no, kind, minimrp_id),
        ).lastrowid


def _setup(db):
    a = _part(db, "ASM-1", "ASSY", 100)
    s = _part(db, "SUB-1", "ASSY", 200)
    c1 = _part(db, "C1", "PART", 1)
    c2 = _part(db, "C2", "PART", 2)
    pm = {100: a, 200: s, 1: c1, 2: c2}
    importer.import_bom_rows(db, parts_map=pm, usedin=[
        {"AutoID": "1", "ParentID": "100", "ChildID": "1", "QtyPer": "2", "RefText": "R1, R2", "LineItemNo": "1"},
        {"AutoID": "2", "ParentID": "100", "ChildID": "200", "QtyPer": "1", "RefText": "", "LineItemNo": "2"},
        {"AutoID": "3", "ParentID": "200", "ChildID": "2", "QtyPer": "3", "RefText": "C5", "LineItemNo": "1"},
    ])
    return a, s, c1, c2


def test_import_and_list(db):
    _setup(db)
    by = {a["part_no"]: a for a in repo.list_assemblies(db)}
    assert set(by) == {"ASM-1", "SUB-1"}
    assert by["ASM-1"]["line_count"] == 2 and by["ASM-1"]["used_in"] == 0
    assert by["SUB-1"]["line_count"] == 1 and by["SUB-1"]["used_in"] == 1  # used in ASM-1


def test_get_assembly_children_subassembly_and_where_used(db):
    a, s, c1, c2 = _setup(db)
    top = repo.get_assembly(db, a)
    kinds = {ln["child_part_no"]: ln["child_kind"] for ln in top["lines"]}
    assert kinds == {"C1": "PART", "SUB-1": "ASSY"}  # a subassembly child
    c1line = next(ln for ln in top["lines"] if ln["child_part_no"] == "C1")
    assert c1line["qty_per"] == 2 and c1line["refdes"] == "R1, R2"
    assert top["used_in"] == []

    sub = repo.get_assembly(db, s)
    assert [ln["child_part_no"] for ln in sub["lines"]] == ["C2"]
    assert [u["part_no"] for u in sub["used_in"]] == ["ASM-1"]  # where-used


def test_summary_and_idempotent_import(db):
    a, s, c1, c2 = _setup(db)
    assert repo.summary(db) == {"assemblies": 2, "lines": 3, "empty": 0}
    importer.import_bom_rows(db, parts_map={100: a, 1: c1}, usedin=[
        {"AutoID": "1", "ParentID": "100", "ChildID": "1", "QtyPer": "2", "LineItemNo": "1"},
    ])
    with db.connect() as conn:
        assert conn.execute("SELECT COUNT(*) FROM bom_lines").fetchone()[0] == 3  # no duplicate


def test_unmapped_rows_are_skipped(db):
    a = _part(db, "ASM-1", "ASSY", 100)
    stats = importer.import_bom_rows(db, parts_map={100: a}, usedin=[
        {"AutoID": "1", "ParentID": "100", "ChildID": "999", "QtyPer": "1"},  # child not in map
    ])
    assert stats == {"bom_lines": 0, "skipped": 1}


def test_get_assembly_rejects_non_assembly(db):
    a, s, c1, c2 = _setup(db)
    assert repo.get_assembly(db, c1) is None  # C1 is a PART, not an assembly


def test_create_assembly(db):
    new_id = repo.create_assembly(
        db, {"part_no": "98-NEW-1", "value": "New product", "rev": "A", "category": "PRODUCT"}
    )
    a = repo.get_assembly(db, new_id)
    assert a is not None and a["part_no"] == "98-NEW-1" and a["kind"] == "ASSY"
    assert a["lines"] == []  # no BOM yet
    assert any(x["part_no"] == "98-NEW-1" for x in repo.list_assemblies(db))


def test_update_assembly_fields(db):
    aid = repo.create_assembly(db, {"part_no": "A1", "value": "v", "rev": "A"})
    repo.update_assembly(db, aid, {"part_no": "A1-REV", "value": "new", "rev": "B",
                                   "category": "PRODUCT", "description": "desc"})
    a = repo.get_assembly(db, aid)
    assert a["part_no"] == "A1-REV" and a["rev"] == "B" and a["value"] == "new"
    assert a["category"] == "PRODUCT" and a["description"] == "desc"
    assert a["kind"] == "ASSY"  # still an assembly


def test_add_and_delete_bom_line(db):
    a, s, c1, c2 = _setup(db)  # ASM-1 already has lines 1 (C1) and 2 (SUB-1)
    repo.add_bom_line(db, a, c2, 4, "U1")
    new = next(ln for ln in repo.get_assembly(db, a)["lines"] if ln["child_part_no"] == "C2")
    assert new["qty_per"] == 4 and new["refdes"] == "U1"
    assert new["line_no"] == 3  # auto-assigned max+1

    repo.delete_bom_line(db, a, new["id"])
    assert all(ln["child_part_no"] != "C2" for ln in repo.get_assembly(db, a)["lines"])


def test_add_bom_line_rejects_self_reference(db):
    a, s, c1, c2 = _setup(db)
    with pytest.raises(ValueError):
        repo.add_bom_line(db, a, a, 1, None)


def test_parts_for_picker_excludes_self(db):
    a, s, c1, c2 = _setup(db)
    ids = {p["id"] for p in repo.parts_for_picker(db, a)}
    assert a not in ids and {s, c1, c2} <= ids


def _component(db, part_no, unit_price):
    from digisearch.web.features.catalog import repo as catrepo
    lines = ([{"supplier_name": "DK", "unit_price": unit_price, "reel_qty": 1, "is_default": True}]
             if unit_price is not None else [])
    return catrepo.create_part(db, part={"part_no": part_no}, supplier_lines=lines)


def test_get_assembly_reports_material_and_loaded(db):
    asm = repo.create_assembly(db, {"part_no": "LQ-ASM"})
    repo.add_bom_line(db, asm, _component(db, "LQ-1", 0.5), 3, "R1")   # material 0.5 x3 = 1.5
    a = repo.get_assembly(db, asm)
    assert a["total_cost"] == pytest.approx(1.5)                        # material
    assert a["loaded_total"] == pytest.approx(1.95)                     # 0.5 x overhead 1.30 x3
    assert "quote_total" not in a                                       # no mfg-margin layer
    ln = a["lines"][0]
    assert ln["loaded_unit"] == pytest.approx(0.65) and ln["loaded_line"] == pytest.approx(1.95)


def test_exclude_from_bom_cost_omits_line_from_totals(db):
    from digisearch.web.features.catalog import repo as catrepo

    asm = repo.create_assembly(db, {"part_no": "EX-ASM"})
    repo.add_bom_line(db, asm, _component(db, "R-1", 0.10), 4, "R1")   # 0.40 material, counted
    stencil = catrepo.create_part(
        db, part={"part_no": "STENCIL-1", "exclude_from_bom_cost": 1},
        supplier_lines=[{"supplier_name": "PCBW", "unit_price": 50.0, "reel_qty": 1, "is_default": True}])
    repo.add_bom_line(db, asm, stencil, 1, None)                        # 50.00, excluded

    a = repo.get_assembly(db, asm)
    # Totals count only the resistor; the stencil is left out of material and loaded.
    assert a["total_cost"] == pytest.approx(0.40)
    assert a["loaded_total"] == pytest.approx(0.52)                     # 0.40 × 1.30
    # …but the stencil line is still present with its own cost shown, flagged excluded.
    sl = next(ln for ln in a["lines"] if ln["child_part_no"] == "STENCIL-1")
    assert sl["line_cost"] == pytest.approx(50.0) and sl["exclude_from_bom_cost"] is True

    # The build estimate and the xlsx export exclude it too.
    assert repo.estimate_bom_cost(db, asm, 1)["material_total"] == pytest.approx(0.40)
    assert repo.get_assembly_for_export(db, asm, build_qty=1)["total_cost"] == pytest.approx(0.40)


def test_rolled_cost_at_is_volume_aware(db):
    from digisearch.web.features.catalog import pricing
    from digisearch.web.features.catalog import repo as catrepo

    leaf = catrepo.create_part(db, part={"part_no": "LC-1"}, supplier_lines=[{
        "supplier_name": "Digi-Key", "supplier_pno": "LC-1-ND", "unit_price": 0.10, "reel_qty": 1,
        "is_default": True,
        "cost_tiers": [{"break_qty": 1, "unit_price": 0.10}, {"break_qty": 1000, "unit_price": 0.03}]}])
    asm = repo.create_assembly(db, {"part_no": "LC-ASM"})
    repo.add_bom_line(db, asm, leaf, 5, "R1")           # 5 of the leaf per board
    with db.connect() as conn:
        assert pricing.rolled_cost_at(conn, asm, 1) == pytest.approx(0.10 * 5)     # 5 leaves -> tier 1
        assert pricing.rolled_cost_at(conn, asm, 200) == pytest.approx(0.03 * 5)   # 1000 -> tier 1000

    # At volume: material (0.03×5) and loaded (×overhead 1.30). No mfg-margin/quote layer.
    est = repo.estimate_bom_cost(db, asm, 200)   # default overhead 1.30
    assert est["build_qty"] == 200
    assert est["material_total"] == pytest.approx(0.15)
    assert est["loaded_total"] == pytest.approx(0.195)     # 0.15 × 1.30
    assert "quote_total" not in est
    line_id = repo.get_assembly(db, asm)["lines"][0]["id"]
    assert est["per_line"][line_id]["material"] == pytest.approx(0.15)
    assert est["per_line"][line_id]["loaded"] == pytest.approx(0.195)


def test_rolled_cost_at_falls_back_to_unit_cost(db):
    from digisearch.web.features.catalog import pricing
    leaf = _component(db, "LC-2", 0.7)                   # flat unit_cost 0.7, no cost tiers
    with db.connect() as conn:
        assert pricing.rolled_cost_at(conn, leaf, 1000) == pytest.approx(0.7)


def test_refresh_bom_for_estimate_skips_in_stock_and_non_distributor(db, monkeypatch):
    from digisearch.models import Candidate
    from digisearch.web.features.catalog import cost_refresh
    from digisearch.web.features.catalog import repo as catrepo

    short = catrepo.create_part(db, part={"part_no": "SH-1"}, supplier_lines=[{
        "supplier_name": "Digi-Key", "supplier_pno": "SH-1-ND", "unit_price": 0.10, "reel_qty": 1,
        "is_default": True, "cost_tiers": [{"break_qty": 1, "unit_price": 0.10}]}], opening={"qty": 0})
    instock = catrepo.create_part(db, part={"part_no": "IN-1"}, supplier_lines=[{
        "supplier_name": "Digi-Key", "supplier_pno": "IN-1-ND", "unit_price": 0.20, "reel_qty": 1,
        "is_default": True, "cost_tiers": [{"break_qty": 1, "unit_price": 0.20}]}], opening={"qty": 1000})
    local = catrepo.create_part(db, part={"part_no": "LO-1"}, supplier_lines=[{
        "supplier_name": "Local Shop", "supplier_pno": "LO-1", "unit_price": 0.5, "reel_qty": 1,
        "is_default": True}], opening={"qty": 0})
    asm = repo.create_assembly(db, {"part_no": "RF-ASM"})
    for leaf in (short, instock, local):
        repo.add_bom_line(db, asm, leaf, 1, None)       # 1 each per board

    class FakeDK:
        def keyword_search(self, kw, limit=5):
            return [Candidate(supplier="Digi-Key", mpn=kw, dk_part_number=kw,
                              price_breaks=[(1, 0.04), (100, 0.02)])]

    monkeypatch.setattr(cost_refresh, "build_clients", lambda: (FakeDK(), None))
    cost_before = {p: catrepo.get_part(db, p)["unit_cost"] for p in (short, instock, local)}

    result = repo.refresh_bom_for_estimate(db, asm, build_qty=10)   # each leaf needs 10
    assert any("SH-1" in m for m in result["refreshed"])            # short distributor -> refreshed
    assert any("IN-1" in m for m in result["in_stock"])            # 1000 on hand >= 10 -> kept
    assert any("LO-1" in m for m in result["skipped"])             # not a distributor -> skipped

    short_cut = sorted((t["break_qty"], t["unit_price"]) for t in
                       catrepo.get_part(db, short)["suppliers"][0]["cost_tiers"] if t["kind"] == "cut")
    assert short_cut == [(1, 0.04), (100, 0.02)]                    # refreshed to the fetched ladder
    instock_cut = [(t["break_qty"], t["unit_price"]) for t in
                   catrepo.get_part(db, instock)["suppliers"][0]["cost_tiers"] if t["kind"] == "cut"]
    assert instock_cut == [(1, 0.20)]                              # in-stock leaf untouched
    for p in (short, instock, local):                              # unit_cost never changes
        assert catrepo.get_part(db, p)["unit_cost"] == cost_before[p]


def test_assembly_line_and_total_costs(db):
    asm = repo.create_assembly(db, {"part_no": "COST-ASM"})
    repo.add_bom_line(db, asm, _component(db, "P1", 0.5), 3, "R1")    # 3 x 0.5 = 1.5
    repo.add_bom_line(db, asm, _component(db, "P2", 2.0), 2, "R2")    # 2 x 2.0 = 4.0
    repo.add_bom_line(db, asm, _component(db, "P3", None), 5, "R3")   # no cost -> blank
    a = repo.get_assembly(db, asm)
    costs = {ln["child_part_no"]: (ln["unit_cost"], ln["line_cost"]) for ln in a["lines"]}
    assert costs["P1"] == (0.5, 1.5)
    assert costs["P2"] == (2.0, 4.0)
    assert costs["P3"][1] is None        # unknown component cost -> no line cost
    assert a["total_cost"] == 5.5        # 1.5 + 4.0


def test_subassembly_cost_rolls_up(db):
    sub = repo.create_assembly(db, {"part_no": "SUB"})
    repo.add_bom_line(db, sub, _component(db, "LEAF", 1.0), 4, "")    # sub = 4 x 1.0 = 4.0
    top = repo.create_assembly(db, {"part_no": "TOP"})
    repo.add_bom_line(db, top, sub, 2, "")                           # 2 subs x 4.0 = 8.0
    a = repo.get_assembly(db, top)
    line = a["lines"][0]
    assert line["unit_cost"] == 4.0 and line["line_cost"] == 8.0
    assert a["total_cost"] == 8.0


# ---- normally_stocked seed (from 90-/98- products) ----

def test_seed_normally_stocked_marks_product_bom_tree(db):
    from digisearch.web.features.assemblies.migrations import SEED_NORMALLY_STOCKED_SQL

    prod90 = repo.create_assembly(db, {"part_no": "90-00999-1"})   # our product
    prod98 = repo.create_assembly(db, {"part_no": "98-00888-1"})   # our product (other level)
    sub = repo.create_assembly(db, {"part_no": "SUB-A"})           # a subassembly under prod90
    leaf1 = _component(db, "LEAF-1", 1.0)                          # direct child of prod90
    leaf2 = _component(db, "LEAF-2", 1.0)                          # nested under sub
    leaf3 = _component(db, "LEAF-3", 1.0)                          # child of prod98
    orphan = _component(db, "ORPHAN", 1.0)                         # not in any 90-/98- tree

    repo.add_bom_line(db, prod90, sub, 1, "")
    repo.add_bom_line(db, prod90, leaf1, 2, "")
    repo.add_bom_line(db, sub, leaf2, 3, "")
    repo.add_bom_line(db, prod98, leaf3, 1, "")

    with db.connect() as conn:
        conn.executescript(SEED_NORMALLY_STOCKED_SQL)
        conn.commit()

    def stocked(pid):
        with db.connect() as conn:
            return bool(conn.execute(
                "SELECT normally_stocked FROM parts WHERE id = ?", (pid,)).fetchone()[0])

    # every descendant of a 90-/98- product (parts AND subassemblies), at any depth
    assert stocked(sub) and stocked(leaf1) and stocked(leaf2) and stocked(leaf3)
    # the products themselves are not marked, and unrelated parts stay off
    assert not stocked(prod90) and not stocked(prod98)
    assert not stocked(orphan)


# ---- xlsx export ----

def test_export_enriches_lines_with_mfr_and_supplier_price(db):
    from digisearch.web.features.assemblies import export_xlsx

    asm = repo.create_assembly(db, {"part_no": "EXP-ASM", "value": "Widget", "rev": "B",
                                    "description": "A test product"})
    p1 = _component(db, "P1", 0.5)      # per-piece price 0.5 -> price_per_uom/qty_per_uom
    with db.connect() as conn:
        conn.execute("UPDATE parts SET mfr_name = 'ACME', mfr_pno = 'ACME-1', "
                     "description = 'A resistor' WHERE id = ?", (p1,))
        conn.commit()
    repo.add_bom_line(db, asm, p1, 3, "R1, R2")

    a = repo.get_assembly_for_export(db, asm, default_markup=1.30)
    ln = a["lines"][0]
    assert ln["child_mfr_name"] == "ACME" and ln["child_mfr_pno"] == "ACME-1"
    assert ln["child_description"] == "A resistor"
    assert ln["child_supplier_price"] == 0.5
    assert ln["unit_cost"] == 0.5 and ln["line_cost"] == 1.5
    assert a["total_cost"] == 1.5
    # loaded = material 0.5 x overhead 1.30 = 0.65 (what the customer pays for the part); line = x3.
    # No manufacturing margin here — profit is on the build, not the parts.
    assert ln["loaded_unit"] == pytest.approx(0.65) and ln["loaded_line"] == pytest.approx(1.95)
    assert a["loaded_total"] == pytest.approx(1.95) and a["build_qty"] == 1

    # The workbook builds and round-trips through openpyxl with the header + total row.
    from io import BytesIO

    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(export_xlsx.workbook_bytes(a)))
    ws = wb["BOM"]
    assert ws["A1"].value == "EXP-ASM  rev B"
    assert ws["A4"].value == "Build volume: 1"
    header = [c.value for c in ws[5]]
    assert header[:3] == ["#", "Qty/board", "Part #"]
    assert "Loaded/unit" in header and "Loaded line" in header
    # data row 6, then total row 7
    row6 = {c.value for c in ws[6]}
    assert "P1" in row6 and "ACME" in row6 and 1.5 in row6
    col_of = {key: i for i, (_n, key, _m) in enumerate(export_xlsx.COLUMNS, start=1)}
    assert ws.cell(row=7, column=col_of["line_cost"]).value == 1.5    # Total material cost
    assert ws.cell(row=7, column=col_of["loaded_line"]).value == pytest.approx(1.95)  # Total loaded


def test_export_sell_price_is_volume_aware(db):
    """A component's loaded cost tier is picked by (qty_per x build volume); the export charges that
    loaded cost (material × overhead) for the parts — no manufacturing margin."""
    from digisearch.web.features.catalog import repo as catrepo

    asm = repo.create_assembly(db, {"part_no": "VOL-ASM"})
    leaf = _component(db, "LEAF", 1.0)
    catrepo.replace_sell_tiers(db, leaf, [{"break_qty": 1, "unit_price": 2.0},
                                          {"break_qty": 1000, "unit_price": 1.0}])
    repo.add_bom_line(db, asm, leaf, 5, "R1")   # 5 of the leaf per board

    # Build 1 -> 5 leaves -> loaded 2.0; line = 2.0*5 = 10.0
    a1 = repo.get_assembly_for_export(db, asm, build_qty=1)
    assert a1["lines"][0]["loaded_unit"] == pytest.approx(2.0)
    assert a1["loaded_total"] == pytest.approx(10.0)

    # Build 200 -> 1000 leaves -> loaded 1.0; line = 1.0*5 = 5.0
    a2 = repo.get_assembly_for_export(db, asm, build_qty=200)
    assert a2["lines"][0]["loaded_unit"] == pytest.approx(1.0)
    assert a2["loaded_total"] == pytest.approx(5.0) and a2["build_qty"] == 200


def test_export_material_uses_cost_tiers_like_page(db):
    """Export material is priced from cost tiers at the build volume (same basis as the BOM page),
    not a flat unit_cost — so it stays proportional to the sell column (overhead × mfg margin)."""
    from digisearch.web.features.catalog import repo as catrepo

    leaf = catrepo.create_part(db, part={"part_no": "EXP-CT"}, supplier_lines=[{
        "supplier_name": "X", "unit_price": 2.0, "reel_qty": 1, "is_default": True,
        "cost_tiers": [{"break_qty": 1, "unit_price": 2.0}, {"break_qty": 100, "unit_price": 1.0}]}])
    asm = repo.create_assembly(db, {"part_no": "EXP-CT-ASM"})
    repo.add_bom_line(db, asm, leaf, 1, "R1")

    a1 = repo.get_assembly_for_export(db, asm, build_qty=1)     # qty 1 -> the 2.0 cost tier
    assert a1["lines"][0]["unit_cost"] == pytest.approx(2.0)
    assert a1["lines"][0]["loaded_unit"] == pytest.approx(2.0 * 1.30)   # material × overhead only
    a100 = repo.get_assembly_for_export(db, asm, build_qty=100)  # qty 100 -> the 1.0 cost tier
    assert a100["lines"][0]["unit_cost"] == pytest.approx(1.0)
    assert a100["lines"][0]["loaded_unit"] == pytest.approx(1.0 * 1.30)


def test_export_returns_none_for_missing_assembly(db):
    assert repo.get_assembly_for_export(db, 99999) is None


def test_export_route_downloads_xlsx(tmp_path):
    from io import BytesIO

    from fastapi.testclient import TestClient
    from openpyxl import load_workbook

    import digisearch.web.app as web_app

    app = web_app.create_app(db_path=tmp_path / "p.db", data_dir=tmp_path / "data",
                             secret_key="test-secret")
    database = app.state.database
    app.state.store.create_user("buyer1", "pw", role="purchasing")

    asm = repo.create_assembly(database, {"part_no": "RT/E-1", "value": "Router"})
    repo.add_bom_line(database, asm, _component(database, "CAP-1", 0.02), 4, "C1")

    client = TestClient(app)
    client.post("/login", data={"username": "buyer1", "password": "pw"}, follow_redirects=False)

    resp = client.get(f"/assemblies/{asm}/export.xlsx")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == \
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # slashes in the part number are sanitised out of the filename
    assert 'filename="RT_E-1_BOM.xlsx"' in resp.headers["content-disposition"]
    wb = load_workbook(BytesIO(resp.content))
    assert "CAP-1" in {c.value for c in wb["BOM"][6]}


# ---- CSV BOM import (reuses purchasing resolution) ----

def _rline(refdes, qty, value, *, chosen=None, status=None, stock_match=None):
    from digisearch.models import BomLine, LineKind, ResolvedLine, Status
    return ResolvedLine(
        line=BomLine(refdes=[refdes], qty=qty, value=value), kind=LineKind.MPN,
        chosen=chosen, status=status or Status.RESOLVED, stock_match=stock_match,
    )


def _cand(mpn, supplier="Digi-Key"):
    from digisearch.models import Candidate
    return Candidate(supplier=supplier, mpn=mpn, manufacturer="ACME",
                     dk_part_number=mpn + "-ND", unit_price=0.1, reel_qty=5000,
                     product_url=f"https://www.digikey.com/{mpn}")


def test_build_import_plan_classifies(db):
    from digisearch.models import Status
    from digisearch.web.features.assemblies import import_bom
    from digisearch.web.features.catalog import repo as catrepo
    from digisearch.web.features.purchasing.service import ResolvedRun

    existing = catrepo.create_part(db, part={"part_no": "EXIST-MPN", "value": "10k"},
                                   supplier_lines=[])
    run = ResolvedRun(
        resolved=[
            _rline("R1", 1, "10k", chosen=_cand("EXIST-MPN")),         # already in catalog
            _rline("R2", 2, "4k7", chosen=_cand("NEW-MPN-1")),         # resolved, new
            _rline("R3", 1, "weird", status=Status.NOT_FOUND),         # unresolved
            _rline("R4", 1, "DNM", status=Status.DNP),                 # skip
            _rline("L1", 1, "", status=Status.MANUAL),                 # under-specified
        ],
        build_qty=1, currency="SEK", stock_checked=False, mouser_enabled=False,
    )
    plan = import_bom.build_import_plan(db, run)
    assert [it["status"] for it in plan] == [
        "in_inventory", "new", "unresolved", "skip", "manual"]
    assert plan[0]["part_id"] == existing
    assert plan[1]["part_no"] == "NEW-MPN-1" and plan[1]["supplier_name"] == "Digi-Key"
    assert plan[1]["product_url"]  # the verify link is carried into the review screen


def test_build_import_plan_builds_ilabs_value(db):
    """A resolved passive gets the iLabs slash notation; a missing field is flagged for review."""
    from digisearch.models import (BomLine, Candidate, CompType, LineKind, PartSpec,
                                   ResolvedLine, Status)
    from digisearch.web.features.assemblies import import_bom
    from digisearch.web.features.purchasing.service import ResolvedRun

    # Resistor: BOM has value+tol+package; power comes from the distributor parametric.
    r_spec = PartSpec(comp_type=CompType.RESISTOR, value_si=1e3, tolerance="5%",
                      package_imperial="0402")
    r_cand = Candidate(supplier="Digi-Key", mpn="R-MPN", dk_part_number="R-MPN-ND",
                       parameters={"Power (Watts)": "0.0625W, 1/16W"})
    # Capacitor: no voltage anywhere -> built partial and flagged.
    c_spec = PartSpec(comp_type=CompType.CAPACITOR, value_si=56e-12, tolerance="5%",
                      package_imperial="0603")
    c_cand = Candidate(supplier="Digi-Key", mpn="C-MPN", dk_part_number="C-MPN-ND", parameters={})
    run = ResolvedRun(
        resolved=[
            ResolvedLine(line=BomLine(refdes=["R1"], qty=1, value="1k"), kind=LineKind.GENERIC_PASSIVE,
                         spec=r_spec, chosen=r_cand, status=Status.RESOLVED),
            ResolvedLine(line=BomLine(refdes=["C1"], qty=1, value="56pF"), kind=LineKind.GENERIC_PASSIVE,
                         spec=c_spec, chosen=c_cand, status=Status.RESOLVED),
        ],
        build_qty=1, currency="SEK", stock_checked=False, mouser_enabled=False,
    )
    plan = import_bom.build_import_plan(db, run)
    assert plan[0]["value"] == "1K/5%/0.0625W/0402" and plan[0]["value_missing"] == []
    assert plan[1]["value"] == "56pF/5%/0603" and plan[1]["value_missing"] == ["voltage"]

    # Applying the flagged part surfaces it in the review list with its part_id.
    asm = repo.create_assembly(db, {"part_no": "ASM-V"})
    result = import_bom.apply_import_plan(db, asm, plan, accepted={0, 1})
    review_pnos = {r["part_no"] for r in result["review"]}
    assert review_pnos == {"C-MPN"}                       # only the cap needs checking
    assert result["review"][0]["missing"] == ["voltage"]


def test_apply_import_plan_creates_and_links(db):
    from digisearch.web.features.assemblies import import_bom
    from digisearch.web.features.catalog import repo as catrepo

    asm = repo.create_assembly(db, {"part_no": "ASM-IMP"})
    existing = catrepo.create_part(db, part={"part_no": "EXIST-MPN"}, supplier_lines=[])
    plan = [
        {"status": "in_inventory", "part_id": existing, "qty": 3, "refdes": "R1"},
        {"status": "new", "part_no": "NEW-1", "value": "4k7", "category": "RESISTOR",
         "mfr_pno": "NEW-1", "supplier_name": "Digi-Key", "supplier_pno": "NEW-1-ND",
         "unit_cost": 0.1, "reel_qty": 5000, "qty": 2, "refdes": "R2"},
        {"status": "unresolved", "part_no": "WEIRD", "value": "weird", "qty": 1, "refdes": "R3"},
        {"status": "new", "part_no": "SKIP-NEW", "qty": 1, "refdes": "R4"},   # not accepted
        {"status": "skip", "qty": 1, "refdes": "R5"},
    ]
    stats = import_bom.apply_import_plan(db, asm, plan, accepted={1, 2})
    assert stats["created"] == 2 and stats["lines_added"] == 3
    childs = {ln["child_part_no"] for ln in repo.get_assembly(db, asm)["lines"]}
    assert childs == {"EXIST-MPN", "NEW-1", "WEIRD"}  # SKIP-NEW left out
    assert catrepo.find_part_id_by_mpn(db, "NEW-1") is not None  # created in catalog


def test_import_captures_supplier_cost_tiers(db):
    """A resolved new part's Digi-Key price breaks become auto-captured cost tiers on create."""
    from digisearch.models import Candidate
    from digisearch.web.features.assemblies import import_bom
    from digisearch.web.features.catalog import repo as catrepo
    from digisearch.web.features.purchasing.service import ResolvedRun

    cand = Candidate(supplier="Digi-Key", mpn="BRK-1", dk_part_number="BRK-1-ND",
                     unit_price=0.10, reel_qty=5000,
                     price_breaks=[(1, 0.10), (100, 0.06), (1000, 0.03)],
                     reel_price_breaks=[(5000, 0.02)])
    run = ResolvedRun(
        resolved=[_rline("R1", 2, "10k", chosen=cand)],
        build_qty=1, currency="SEK", stock_checked=False, mouser_enabled=False,
    )
    asm = repo.create_assembly(db, {"part_no": "ASM-BRK"})
    plan = import_bom.build_import_plan(db, run)
    assert plan[0]["cost_tiers"] == [{"break_qty": 1, "unit_price": 0.10},
                                     {"break_qty": 100, "unit_price": 0.06},
                                     {"break_qty": 1000, "unit_price": 0.03}]
    import_bom.apply_import_plan(db, asm, plan, accepted={0})

    pid = catrepo.find_part_id_by_mpn(db, "BRK-1")
    tiers = catrepo.get_part(db, pid)["suppliers"][0]["cost_tiers"]
    cut = [(t["break_qty"], t["unit_price"]) for t in tiers if t["kind"] == "cut"]
    reel = [(t["break_qty"], t["unit_price"]) for t in tiers if t["kind"] == "reel"]
    assert cut == [(1, 0.10), (100, 0.06), (1000, 0.03)]
    assert reel == [(5000, 0.02)]


def test_convert_to_component_reclassifies_empty_assembly(db):
    from digisearch.web.features.catalog import repo as catrepo

    # A part mis-entered as an assembly, with stock but no BOM.
    aid = repo.create_assembly(db, {"part_no": "WAS-ASSY", "value": "really a part",
                                    "default_build_days": 5})
    with db.connect() as conn:
        conn.execute("UPDATE parts SET total_qty = 7 WHERE id = ?", (aid,))
        conn.commit()

    repo.convert_to_component(db, aid)

    part = catrepo.get_part(db, aid)
    assert part["kind"] == "PART" and part["default_build_days"] is None
    assert part["total_qty"] == 7                       # stock preserved
    assert repo.get_assembly(db, aid) is None           # no longer an assembly
    # shows up in the parts catalog now
    assert any(p["part_no"] == "WAS-ASSY" for p in catrepo.list_parts(db)[0])


def test_convert_blocked_when_bom_present(db):
    a, s, c1, c2 = _setup(db)
    with pytest.raises(ValueError, match="BOM line"):
        repo.convert_to_component(db, a)
    assert repo.get_assembly(db, a) is not None          # untouched


def test_convert_rejects_non_assembly(db):
    from digisearch.web.features.catalog import repo as catrepo
    pid = catrepo.create_part(db, part={"part_no": "PLAIN"}, supplier_lines=[])
    with pytest.raises(ValueError, match="not an assembly"):
        repo.convert_to_component(db, pid)


def test_convert_used_in_links_survive(db):
    # SUB-1 (an ASSY) is used in ASM-1; emptying SUB-1's own BOM lets it become a component
    # while staying a child of ASM-1.
    a, s, c1, c2 = _setup(db)
    for ln in repo.get_assembly(db, s)["lines"]:
        repo.delete_bom_line(db, s, ln["id"])
    repo.convert_to_component(db, s)
    parent = repo.get_assembly(db, a)
    assert "SUB-1" in {ln["child_part_no"] for ln in parent["lines"]}
