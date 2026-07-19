import pytest

from digisearch.web.core import FeatureRegistry
from digisearch.web.core.db import Database
from digisearch.web.features.article_register import feature as article_register_feature
from digisearch.web.features.article_register import importer, repo
from digisearch.web.features.article_register.codes import article_code
from digisearch.web.features.article_register.repo import DuplicateNumber
from digisearch.web.features.catalog import feature as catalog_feature


@pytest.fixture
def db(tmp_path):
    database = Database(tmp_path / "a.db")
    reg = FeatureRegistry()
    # catalog first (as in production) so `parts` exists for the soft code→part_no join.
    reg.register(catalog_feature, article_register_feature)
    database.apply_migrations(reg)
    return database


def test_prefixes_seeded_and_grouped(db):
    prefixes = repo.list_prefixes(db)
    assert len(prefixes) == 25
    by_code = {p["code"]: p for p in prefixes}
    assert by_code["98"]["label"] == "Assemblies" and by_code["98"]["category"] == "internal"
    assert by_code["05"]["label"] == "Brainlit" and by_code["05"]["category"] == "customer"
    groups = {g["category"]: len(g["prefixes"]) for g in repo.prefixes_grouped(db)}
    assert groups == {"customer": 8, "ic": 1, "document": 10, "internal": 6}


def test_allocation_and_code_format(db):
    assert repo.next_running_no(db) == 1
    eid = repo.create_entry(db, prefix="98", running_no=2, suffix=4, product="Enterprise PCBA")
    entry = repo.get_entry(db, eid)
    assert entry["code"] == "98-00002-4" == article_code("98", 2, 4)
    assert repo.next_running_no(db) == 1  # 1 is free — reuse the gap, not MAX+1
    assert repo.next_suffix(db, "98", 2) == 5  # MAX(4)+1 within the family
    assert repo.next_suffix(db, "99", 2) == 1  # different prefix, empty


def test_next_running_no_fills_the_first_gap(db):
    for n in (1, 2, 3, 100, 101):  # a legacy block + a far-off high block, gap at 4
        repo.create_entry(db, prefix="98", running_no=n, suffix=1)
    assert repo.next_running_no(db) == 4  # earliest free number, not MAX(101)+1
    repo.create_entry(db, prefix="98", running_no=4, suffix=1)
    assert repo.next_running_no(db) == 5

    # Retired and reserved rows still occupy their running number — never reused.
    rid = repo.create_entry(db, prefix="99", running_no=5, suffix=1)
    repo.set_retired(db, rid, True)
    with db.connect() as conn:  # a reserved (prefix/suffix/code NULL) running number
        conn.execute("INSERT INTO article_numbers (running_no) VALUES (6)")
        conn.commit()
    assert repo.next_running_no(db) == 7


def test_prefix_is_zero_padded(db):
    eid = repo.create_entry(db, prefix="5", running_no=357, suffix=1)  # customer 5 → '05'
    assert repo.get_entry(db, eid)["code"] == "05-00357-1"


def test_apply_template_skips_codes_already_in_family(db):
    tid = repo.create_template(db, name="Full")
    repo.save_template(db, tid, name="Full", notes=None, lines=[
        {"prefix": "98", "suffix": 1, "label": ""},
        {"prefix": "99", "suffix": 1, "label": "PCBA"},
        {"prefix": "54", "suffix": 1, "label": "drawing"},
    ])
    rn = repo.apply_template(db, tid, product="Widget")  # new family — all three created
    assert repo.family_prefixes(db, rn) == ["54", "98", "99"]

    # A second template shares the exact code 98-…-1 (already present) and adds a new 51 group.
    tid2 = repo.create_template(db, name="Extra")
    repo.save_template(db, tid2, name="Extra", notes=None, lines=[
        {"prefix": "98", "suffix": 1, "label": ""},
        {"prefix": "51", "suffix": 1, "label": "stencil"},
    ])
    repo.apply_template(db, tid2, product="Widget", running_no=rn)
    fam = repo.get_family(db, rn)
    assert sum(1 for e in fam if e["prefix"] == "98") == 1  # 98-…-1 not duplicated
    assert article_code("51", rn, 1) in {e["code"] for e in fam}  # 51 added

    # Every code of the first template is now present → nothing to add.
    with pytest.raises(ValueError, match="nothing to add"):
        repo.apply_template(db, tid, product="Widget", running_no=rn)


def test_apply_template_adds_new_suffixes_under_an_existing_prefix(db):
    # Regression: a template with several lines sharing a prefix (99: PCB/Stencil TOP/Stencil BOT)
    # must add the missing suffixes even though the prefix 99 is already in the family — the skip is
    # by exact code, not by whole prefix.
    tid = repo.create_template(db, name="PCB")
    repo.save_template(db, tid, name="PCB", notes=None, lines=[
        {"prefix": "98", "suffix": 1, "label": ""},
        {"prefix": "99", "suffix": 1, "label": "PCB"},
        {"prefix": "99", "suffix": 2, "label": "Stencil TOP"},
        {"prefix": "99", "suffix": 3, "label": "Stencil BOT"},
    ])
    # Start a family that only has the assembly + the PCB (98-…-1, 99-…-1), as if created earlier.
    rn = repo.next_running_no(db)
    repo.create_entry(db, prefix="98", running_no=rn, suffix=1, product="MB")
    repo.create_entry(db, prefix="99", running_no=rn, suffix=1, product="MB - PCB")

    repo.apply_template(db, tid, product="MB", running_no=rn)
    codes = {e["code"] for e in repo.get_family(db, rn)}
    assert article_code("99", rn, 2) in codes  # Stencil TOP created
    assert article_code("99", rn, 3) in codes  # Stencil BOT created
    assert sum(1 for e in repo.get_family(db, rn) if e["prefix"] == "99") == 3  # 99-1 not duplicated


def test_duplicate_triplet_rejected(db):
    repo.create_entry(db, prefix="98", running_no=390, suffix=1)
    with pytest.raises(DuplicateNumber):
        repo.create_entry(db, prefix="98", running_no=390, suffix=1)


def test_create_product_makes_one_family(db):
    # A product = one running number with a line per ticked group, all suffix 1.
    running_no = repo.create_product(db, product="AddBox 200", prefixes=["98", "54", "99"],
                                     created_by="LO")
    assert running_no == 1
    family = repo.get_family(db, running_no)
    codes = [e["code"] for e in family]
    assert codes == ["54-00001-1", "98-00001-1", "99-00001-1"]  # ordered by prefix
    assert all(e["product"] == "AddBox 200" and e["created_by"] == "LO" for e in family)
    # Next product gets the next running number.
    assert repo.create_product(db, product="AddBox 300", prefixes=["98"]) == 2


def test_create_product_dedupes_and_requires_a_group(db):
    running_no = repo.create_product(db, product="X", prefixes=["98", "98", ""])
    assert len(repo.get_family(db, running_no)) == 1  # duplicate 98 collapsed to one line
    with pytest.raises(ValueError):
        repo.create_product(db, product="Y", prefixes=[])


def test_retire_flag_does_not_block_other_suffix(db):
    a = repo.create_entry(db, prefix="99", running_no=2, suffix=4, product="V1.2 PCB")
    repo.set_retired(db, a, True)
    assert repo.get_entry(db, a)["retired"] == 1
    # A different suffix in the same family is still allocatable.
    b = repo.create_entry(db, prefix="99", running_no=2, suffix=5)
    assert repo.get_entry(db, b)["code"] == "99-00002-5"
    repo.set_retired(db, a, False)
    assert repo.get_entry(db, a)["retired"] == 0


def test_family_view_gathers_the_running_number(db):
    repo.create_entry(db, prefix="97", running_no=2, suffix=4)
    repo.create_entry(db, prefix="98", running_no=2, suffix=4)
    repo.create_entry(db, prefix="99", running_no=2, suffix=1)
    repo.create_entry(db, prefix="98", running_no=3, suffix=1)  # different family
    codes = [e["code"] for e in repo.get_family(db, 2)]
    assert codes == ["97-00002-4", "98-00002-4", "99-00002-1"]


def test_soft_catalog_link(db):
    with db.connect() as conn:
        conn.execute("INSERT INTO parts (part_no, kind) VALUES ('98-00002-4', 'ASSY')")
        conn.commit()
    eid = repo.create_entry(db, prefix="98", running_no=2, suffix=4)
    entry = next(e for e in repo.list_entries(db) if e["id"] == eid)
    assert entry["part_id"] is not None
    # A code with no matching part has a null link.
    other = repo.create_entry(db, prefix="99", running_no=2, suffix=1)
    assert next(e for e in repo.list_entries(db) if e["id"] == other)["part_id"] is None


def test_search_unassigned_excludes_associated_retired_and_reserved(db):
    free_assy = repo.create_entry(db, prefix="98", running_no=2, suffix=1, product="Gateway board")
    taken = repo.create_entry(db, prefix="98", running_no=3, suffix=1, product="Router board")
    retired = repo.create_entry(db, prefix="98", running_no=4, suffix=1, product="Old board")
    repo.set_retired(db, retired, True)
    free_comp = repo.create_entry(db, prefix="99", running_no=2, suffix=1, product="Widget")
    repo.create_entry(db, prefix=None, running_no=99, suffix=None)  # reserved: no code
    with db.connect() as conn:  # associate one code with a catalog part/assembly
        conn.execute("INSERT INTO parts (part_no, kind) VALUES ('98-00003-1', 'ASSY')")
        conn.commit()

    all_codes = {r["code"] for r in repo.search_unassigned(db)}
    assert "98-00002-1" in all_codes and "99-00002-1" in all_codes
    assert "98-00003-1" not in all_codes  # already has a part
    assert "98-00004-1" not in all_codes  # retired
    assert None not in all_codes          # reserved row (no code) excluded
    _ = (free_assy, taken, retired, free_comp)

    # prefix scoping (e.g. assemblies only) and text match on code/product.
    scoped = {r["code"] for r in repo.search_unassigned(db, prefix="98")}
    assert scoped == {"98-00002-1"}
    assert {r["code"] for r in repo.search_unassigned(db, "gateway")} == {"98-00002-1"}
    assert {r["code"] for r in repo.search_unassigned(db, "99-0000")} == {"99-00002-1"}


def test_list_filters_and_retired_visibility(db):
    good = repo.create_entry(db, prefix="98", running_no=2, suffix=4, product="Enterprise")
    dead = repo.create_entry(db, prefix="99", running_no=2, suffix=4, product="Old PCB")
    repo.set_retired(db, dead, True)
    active = {e["id"] for e in repo.list_entries(db)}
    assert good in active and dead not in active  # retired hidden by default
    assert dead in {e["id"] for e in repo.list_entries(db, include_retired=True)}
    assert {e["id"] for e in repo.list_entries(db, category="internal")} == {good}
    assert {e["id"] for e in repo.list_entries(db, search="enterprise")} == {good}


def _client(tmp_path):
    from fastapi.testclient import TestClient

    import digisearch.web.app as web_app
    app = web_app.create_app(db_path=tmp_path / "p.db", data_dir=tmp_path / "data",
                             secret_key="test-secret")
    app.state.store.create_user("buyer1", "pw", role="purchasing")
    client = TestClient(app)
    client.post("/login", data={"username": "buyer1", "password": "pw"}, follow_redirects=False)
    return app, client


def test_allocate_returns_to_create_page_with_code(tmp_path):
    app, client = _client(tmp_path)
    db = app.state.database

    # New Number form carries the preset prefix + hidden return path.
    page = client.get("/article-register/new", params={"prefix": "99", "return_to": "/catalog/new"})
    assert 'name="return_to" value="/catalog/new"' in page.text
    assert 'value="99" selected' in page.text

    nn = repo.next_running_no(db)
    r = client.post("/article-register/new",
                    data={"mode": "new", "prefix": "99", "return_to": "/catalog/new"},
                    follow_redirects=False)
    assert r.headers["location"] == f"/catalog/new?part_no=99-{nn:05d}-1"
    # …and the create-part page prefills that number.
    assert f'value="99-{nn:05d}-1"' in client.get(r.headers["location"]).text

    # New Product returns the assembly (98) code to the New-assembly page.
    rn = repo.next_running_no(db)
    r = client.post("/article-register/product",
                    data={"product": "GW", "prefixes": ["98", "99", "54"],
                          "return_to": "/assemblies/new"}, follow_redirects=False)
    assert r.headers["location"] == f"/assemblies/new?part_no=98-{rn:05d}-1"
    assert f'value="98-{rn:05d}-1"' in client.get(r.headers["location"]).text


def test_from_template_returns_assembly_code(tmp_path):
    app, client = _client(tmp_path)
    db = app.state.database
    # The seeded 'Standard PCB product' template (id 1) includes a 98 assembly line.
    page = client.get("/article-register/from-template", params={"return_to": "/assemblies/new"})
    assert 'name="return_to" value="/assemblies/new"' in page.text

    from urllib.parse import parse_qs, urlparse

    rn = repo.next_running_no(db)
    r = client.post("/article-register/from-template",
                    data={"template_id": "1", "product": "GW mobo", "mode": "new",
                          "return_to": "/assemblies/new"}, follow_redirects=False)
    loc = r.headers["location"]
    assert urlparse(loc).path == "/assemblies/new"
    assert parse_qs(urlparse(loc).query)["part_no"][0] == f"98-{rn:05d}-1"
    assert f'value="98-{rn:05d}-1"' in client.get(loc).text


def test_apply_template_link_carries_family_product(tmp_path):
    app, client = _client(tmp_path)
    db = app.state.database
    # A family whose assembly (98) line carries the product name (template 1 = Standard PCB product).
    rn = repo.apply_template(db, 1, product="MiThings GW")

    # The detail page's "Apply template" link carries the running number AND the product.
    page = client.get(f"/article-register/{rn}").text
    assert f"running_no={rn}" in page
    assert ("product=MiThings%20GW" in page) or ("product=MiThings+GW" in page)

    # Following it lands on the dialog with the product prefilled, in existing-family mode.
    form = client.get("/article-register/from-template",
                      params={"running_no": rn, "product": "MiThings GW"}).text
    assert 'value="MiThings GW"' in form
    assert 'name="mode" value="existing"' in form


def test_from_template_creates_stub_parts_and_dialog(tmp_path):
    from urllib.parse import parse_qs, urlparse

    from digisearch.web.features.catalog import repo as crepo
    from digisearch.web.features.documents import repo as drepo

    app, client = _client(tmp_path)
    db = app.state.database
    rn = repo.next_running_no(db)
    r = client.post("/article-register/from-template",
                    data={"template_id": "1", "product": "MiThings GW", "mode": "new",
                          "return_to": "/assemblies/new"}, follow_redirects=False)
    qs = parse_qs(urlparse(r.headers["location"]).query)
    assert qs["part_no"][0] == f"98-{rn:05d}-1"
    assert qs["desc"][0] == "MiThings GW"
    created = qs["created"][0].split(",")

    # The component (99) lines become kind=PART stubs; the 98 assembly does not.
    assert f"98-{rn:05d}-1" not in created
    assert len(created) == 3  # 3×99 (the 3×54 document lines become documents, not parts)
    for code in created:
        assert code.startswith(f"99-{rn:05d}-")
        p = crepo.find_part_by_part_no(db, code)
        # house convention: the product/description lands in `value`, not `description`
        assert p and p["kind"] == "PART" and p["value"] and not p["description"]

    # The 54 document lines become document items (not parts).
    for i in (1, 2, 3):
        doc_code = f"54-{rn:05d}-{i}"
        assert doc_code not in created and crepo.find_part_by_part_no(db, doc_code) is None
        doc = drepo.document_for_code(db, doc_code)
        assert doc is not None and doc["storage_kind"] == "file"

    # The return page prefills part_no + name and raises the "created, edit before use" dialog.
    page = client.get(r.headers["location"]).text
    assert f'value="98-{rn:05d}-1"' in page and 'value="MiThings GW"' in page
    assert "parts created" in page and created[0] in page

    # Re-running the same template into the *same* family adds nothing: every group is already
    # present, so all lines are skipped and the form re-renders with an explanatory error (no
    # redirect, no new/duplicate parts).
    r2 = client.post("/article-register/from-template",
                     data={"template_id": "1", "product": "MiThings GW", "mode": "existing",
                           "running_no": str(rn), "return_to": "/assemblies/new"},
                     follow_redirects=False)
    assert r2.status_code == 400 and "location" not in r2.headers
    assert "nothing to add" in r2.text
    for code in created:
        assert crepo.find_part_by_part_no(db, code)  # originals untouched, exactly one each


def test_from_template_parts_land_in_new_assembly_bom(tmp_path):
    from urllib.parse import parse_qs, urlparse

    app, client = _client(tmp_path)
    db = app.state.database
    r = client.post("/article-register/from-template",
                    data={"template_id": "1", "product": "GW", "mode": "new",
                          "return_to": "/assemblies/new"}, follow_redirects=False)
    q = parse_qs(urlparse(r.headers["location"]).query)
    created = q["created"][0].split(",")

    # Creating the assembly (carrying the hidden bom_parts) adds every created part to its BOM at qty 1.
    r2 = client.post("/assemblies/new",
                     data={"part_no": q["part_no"][0], "value": "GW", "bom_parts": ",".join(created)},
                     follow_redirects=False)
    asm_id = int(r2.headers["location"].rsplit("/", 1)[1])
    with db.connect() as conn:
        rows = [dict(x) for x in conn.execute(
            "SELECT c.part_no, b.qty_per FROM bom_lines b JOIN parts c ON c.id = b.child_id "
            "WHERE b.parent_id = ? ORDER BY b.line_no", (asm_id,))]
    assert {x["part_no"] for x in rows} == set(created)
    assert all(x["qty_per"] == 1 for x in rows)


def test_allocate_rejects_offsite_return_to(tmp_path):
    _, client = _client(tmp_path)
    # An external return path is ignored; allocation falls back to the family detail page.
    r = client.post("/article-register/new",
                    data={"mode": "new", "prefix": "99", "return_to": "https://evil.example"},
                    follow_redirects=False)
    assert r.headers["location"].startswith("/article-register/")
    # Protocol-relative is also rejected.
    r = client.post("/article-register/new",
                    data={"mode": "new", "prefix": "99", "return_to": "//evil.example"},
                    follow_redirects=False)
    assert r.headers["location"].startswith("/article-register/")


def _build_workbook(path):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Artikellista"
    ws.append(["Prefix", "Artikelnummer", "Suffix", "Produkt", "Upplagd av", "Shortform"])
    data = [
        ("98", 2, 4, "Enterprise PCBA", "PO", "98-00002-4", False),
        ("99", 2, 4, "Enterprise PCB (old)", "PO", "99-00002-4", True),  # struck → retired
        ("", 50, "", "", "", "", False),                                 # reserved running number
        ("", "", "", "", "", "", False),                                 # empty → skipped
    ]
    for prefix, num, suffix, product, by, code, struck in data:
        row = ws.max_row + 1
        ws.cell(row=row, column=1, value=prefix)
        ws.cell(row=row, column=2, value=num)
        ws.cell(row=row, column=3, value=suffix)
        ws.cell(row=row, column=4, value=product)
        ws.cell(row=row, column=5, value=by)
        code_cell = ws.cell(row=row, column=6, value=code)
        if struck:
            code_cell.font = Font(strike=True)
    wb.save(path)


def test_importer_marks_retired_and_skips_reserved(db, tmp_path):
    xlsx = tmp_path / "reg.xlsx"
    _build_workbook(xlsx)
    stats = importer.import_register(db, xlsx)
    assert stats["article numbers"] == 2
    assert stats["reserved skipped"] == 1  # blank-prefix placeholder row is NOT imported
    assert stats["article retired"] == 1
    # Only the two assigned numbers land; no reserved rows.
    assert repo.summary(db) == {"total": 2, "reserved": 0, "retired": 1, "families": 1}

    fam = {e["code"]: e for e in repo.get_family(db, 2)}
    assert fam["99-00002-4"]["retired"] == 1
    assert fam["98-00002-4"]["retired"] == 0
    assert repo.get_family(db, 50) == []  # the reserved running number was skipped


def test_importer_is_idempotent(db, tmp_path):
    xlsx = tmp_path / "reg.xlsx"
    _build_workbook(xlsx)
    importer.import_register(db, xlsx)
    before = repo.summary(db)
    stats = importer.import_register(db, xlsx)
    # Re-run: both assigned rows already exist (skipped) + the empty row = 3 skipped.
    assert stats == {"article numbers": 0, "article retired": 0,
                     "reserved skipped": 1, "article skipped": 3}
    assert repo.summary(db) == before
