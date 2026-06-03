"""Write resolved BOM lines to a rich Excel workbook with a cost summary."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..models import ResolvedLine, Status

HEADERS = [
    "RefDes", "Qty/board", "Total qty",
    "In stock (free)", "Need to buy", "miniMRP match",
    "Original value", "Device", "Package", "Kind", "Status", "Confidence",
    "Supplier", "Manufacturer", "MPN", "Supplier #", "Lifecycle", "Supplier stock",
    "Unit price (q1)", "Packaging", "Order qty", "Order unit price", "Line cost",
    "Datasheet", "Flag", "Alternates",
]
_STATUS_COL = HEADERS.index("Status") + 1
_MONEY_COLS = [HEADERS.index(h) + 1 for h in ("Unit price (q1)", "Order unit price", "Line cost")]

_STATUS_FILL = {
    Status.RESOLVED: "C6EFCE",   # green
    Status.REVIEW: "FFEB9C",     # amber
    Status.IN_STOCK: "BDD7EE",   # blue — already on the shelf
    Status.NOT_FOUND: "FFC7CE",  # red
    Status.ERROR: "FFC7CE",
    Status.DNP: "D9D9D9",        # grey
    Status.NON_ORDERABLE: "D9D9D9",
}
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")


def _alternates_text(line: ResolvedLine) -> str:
    return " | ".join(
        f"{c.mpn} ({c.dk_part_number})" for c in line.alternates if c.mpn
    )


def _row_for(line: ResolvedLine, build_qty: int) -> tuple[list, float | None]:
    total_qty = line.line.qty * build_qty
    c = line.chosen
    row = [
        line.line.refdes_str,
        line.line.qty,
        total_qty,
        int(line.stock_free) if line.stock_free is not None else None,
        line.need_to_buy,
        line.stock_match,
        line.line.value or line.line.device,
        line.line.device,
        line.line.package,
        line.kind.value,
        line.status.value,
        round(line.confidence, 3) if c else None,
        c.supplier if c else None,
        c.manufacturer if c else None,
        c.mpn if c else None,
        c.dk_part_number if c else None,
        c.lifecycle if c else None,
        c.quantity_available if c else None,
        line.unit_price(),
        line.packaging,
        line.purchase_qty,
        line.purchase_unit_price,
        line.line_cost,
        c.datasheet_url if c else None,
        line.flag_reason,
        _alternates_text(line),
    ]
    return row, line.line_cost


def write_report(
    lines: list[ResolvedLine],
    path: str | Path,
    build_qty: int = 1,
    currency: str = "",
) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"

    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"

    total_cost = 0.0
    for r, line in enumerate(lines, start=2):
        row, ext_cost = _row_for(line, build_qty)
        for col, value in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=value)
        fill = _STATUS_FILL.get(line.status)
        if fill:
            ws.cell(row=r, column=_STATUS_COL).fill = PatternFill("solid", fgColor=fill)
        if ext_cost:
            total_cost += ext_cost

    # currency number format on price columns
    money_fmt = f'#,##0.0000 "{currency}"' if currency else "#,##0.0000"
    for col in _MONEY_COLS:
        for r in range(2, len(lines) + 2):
            ws.cell(row=r, column=col).number_format = money_fmt

    _autosize(ws)
    _write_summary(wb, lines, build_qty, total_cost, currency)

    path = Path(path)
    wb.save(path)
    return path


def _write_summary(wb, lines, build_qty, total_cost, currency):
    ws = wb.create_sheet("Summary")
    counts: dict[str, int] = {}
    for line in lines:
        counts[line.status.value] = counts.get(line.status.value, 0) + 1

    rows = [
        ("Build quantity (boards)", build_qty),
        ("BOM lines", len(lines)),
        ("Unique parts to order", sum(1 for l in lines if l.chosen)),
        ("", ""),
        ("Total BOM cost @build", round(total_cost, 4)),
        ("Cost per board", round(total_cost / build_qty, 4) if build_qty else 0),
        ("", ""),
    ]
    for status, n in sorted(counts.items()):
        rows.append((f"Lines: {status}", n))

    ws.cell(row=1, column=1, value="DigiSearch resolution summary").font = Font(bold=True, size=14)
    for i, (label, value) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        ws.cell(row=i, column=2, value=value)
    if currency:
        for i, (label, _) in enumerate(rows, start=3):
            if "cost" in label.lower():
                ws.cell(row=i, column=2).number_format = f'#,##0.00 "{currency}"'

    # Flagged / unresolved detail
    start = len(rows) + 5
    ws.cell(row=start, column=1, value="Needs attention").font = Font(bold=True, size=12)
    ws.cell(row=start + 1, column=1, value="RefDes").font = Font(bold=True)
    ws.cell(row=start + 1, column=2, value="Original").font = Font(bold=True)
    ws.cell(row=start + 1, column=3, value="Status").font = Font(bold=True)
    ws.cell(row=start + 1, column=4, value="Reason").font = Font(bold=True)
    r = start + 2
    for line in lines:
        if line.flagged:
            ws.cell(row=r, column=1, value=line.line.refdes_str)
            ws.cell(row=r, column=2, value=line.line.value or line.line.device)
            ws.cell(row=r, column=3, value=line.status.value)
            ws.cell(row=r, column=4, value=line.flag_reason)
            r += 1
    _autosize(ws)


def _autosize(ws, max_width: int = 60):
    for col_cells in ws.columns:
        width = max(
            (len(str(c.value)) for c in col_cells if c.value is not None), default=10
        )
        letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[letter].width = min(max_width, max(10, width + 2))
