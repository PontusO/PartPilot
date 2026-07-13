"""One-time seed of the Article Register from the legacy ``Artikelregister`` Excel workbook.

The ``Artikellista`` sheet has columns: A=prefix, B=running number, C=suffix, D=product,
E=created-by initials, F=assembled code (a formula we ignore — we recompute the code ourselves).
A struck-through code cell (column F) marks a retired number that must never be reused. Blank-prefix
rows are the workbook's pre-reserved placeholders — we do NOT import them; numbers are allocated on
demand in the app (``repo.create_product`` / the single-number allocator) as the need arises.

Idempotent: re-running skips rows already present (assigned rows dedupe on the unique
code/triplet indexes), so it is safe during the dual-run period while miniMRP/Excel remain the
historical record.
"""

from __future__ import annotations

from pathlib import Path

from ...core.db import Database
from .codes import article_code, normalize_prefix

SHEET = "Artikellista"


def import_register(db: Database, xlsx_path: str | Path) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, read_only=True)  # values for A–E, strike font on F
    if SHEET not in wb.sheetnames:
        raise ValueError(f"'{SHEET}' sheet not found in {xlsx_path}")
    ws = wb[SHEET]

    stats = {"assigned": 0, "retired": 0, "reserved_skipped": 0, "skipped": 0}
    with db.connect() as conn:
        for row in ws.iter_rows(min_row=2):
            prefix = normalize_prefix(row[0].value)
            running_no = _int(row[1].value)
            suffix = _int(row[2].value)
            product = _clean(row[3].value)
            created_by = _clean(row[4].value)
            code_cell = row[5] if len(row) > 5 else None
            retired = bool(code_cell is not None and code_cell.font and code_cell.font.strike)

            if running_no is None:
                stats["skipped"] += 1
                continue

            if prefix is None or suffix is None:
                # Blank-prefix rows are the workbook's pre-reserved placeholders. We do NOT import
                # them: numbers are allocated on demand in the app as the need comes up, so
                # pre-seeding hundreds of empty rows just clutters the register.
                stats["reserved_skipped"] += 1
                continue

            code = article_code(prefix, running_no, suffix)
            cur = conn.execute(
                """INSERT OR IGNORE INTO article_numbers
                       (prefix, running_no, suffix, code, product, created_by, retired, source)
                   VALUES (?, ?, ?, ?, ?, ?, ?, 'excel')""",
                (prefix, running_no, suffix, code, product, created_by, 1 if retired else 0))
            if cur.rowcount:
                stats["assigned"] += 1
                if retired:
                    stats["retired"] += 1
            else:
                stats["skipped"] += 1
        conn.commit()
    wb.close()
    return {
        "article numbers": stats["assigned"],
        "article retired": stats["retired"],
        "reserved skipped": stats["reserved_skipped"],
        "article skipped": stats["skipped"],
    }


def _int(value) -> int | None:
    try:
        text = str(value).strip()
        return int(float(text)) if text else None
    except (TypeError, ValueError):
        return None


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None
