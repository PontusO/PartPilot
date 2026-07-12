"""Render a product/assembly BOM to a customer-facing Excel workbook.

Kept deliberately small and independent of the CLI's ``report/excel.py`` (which is bound to
the ``ResolvedLine`` model). This one works off the plain dicts from
``repo.get_assembly_for_export`` so the assemblies feature owns its own export shape.
"""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# (header, dict key, is_money) — column order for the BOM table.
COLUMNS = [
    ("#", "line_no", False),
    ("Qty/board", "qty_per", False),
    ("Part #", "child_part_no", False),
    ("Manufacturer", "child_mfr_name", False),
    ("Mfr P/N", "child_mfr_pno", False),
    ("Description", "child_description", False),
    ("Value", "child_value", False),
    ("Category", "child_category", False),
    ("Supplier price", "child_supplier_price", True),
    ("Unit cost", "unit_cost", True),
    ("Line cost", "line_cost", True),
    ("Sell/unit", "sell_unit", True),
    ("Sell line", "sell_line", True),
    ("Reference designators", "refdes", False),
]
# The two total-bearing columns (cost and sell) get a totals row at the bottom.
_TOTAL_COLS = [("line_cost", "total_cost"), ("sell_line", "sell_total")]

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_TOTAL_FILL = PatternFill("solid", fgColor="DDEBF7")


def build_workbook(assembly: dict, currency: str = "SEK") -> Workbook:
    """Build the workbook for one assembly (as returned by ``get_assembly_for_export``)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "BOM"
    money_fmt = f'#,##0.0000 "{currency}"' if currency else "#,##0.0000"

    # --- product header block ---
    title = assembly["part_no"]
    if assembly.get("rev"):
        title += f"  rev {assembly['rev']}"
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    subtitle = assembly.get("value") or ""
    if assembly.get("description"):
        subtitle = f"{subtitle} — {assembly['description']}" if subtitle else assembly["description"]
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = Font(italic=True)
    ws.cell(row=3, column=1, value=f"BOM lines: {len(assembly['lines'])}").font = Font(bold=True)
    ws.cell(row=4, column=1,
            value=f"Build volume: {assembly.get('build_qty', 1)}").font = Font(bold=True)

    # --- table header ---
    header_row = 5
    for col, (name, _key, _money) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    # --- lines ---
    money_cols = [i for i, (_n, _k, m) in enumerate(COLUMNS, start=1) if m]
    r = header_row + 1
    for idx, ln in enumerate(assembly["lines"], start=1):
        for col, (_name, key, _money) in enumerate(COLUMNS, start=1):
            value = ln.get(key)
            if key == "line_no" and value is None:
                value = idx
            ws.cell(row=r, column=col, value=value)
        for col in money_cols:
            ws.cell(row=r, column=col).number_format = money_fmt
        r += 1

    # --- total row (a total under each of the cost and sell columns) ---
    col_of = {key: i for i, (_n, key, _m) in enumerate(COLUMNS, start=1)}
    first_total_col = min(col_of[k] for k, _ in _TOTAL_COLS)
    label_cell = ws.cell(row=r, column=first_total_col - 1, value="Total")
    label_cell.font = Font(bold=True)
    label_cell.alignment = Alignment(horizontal="right")
    for line_key, total_key in _TOTAL_COLS:
        cell = ws.cell(row=r, column=col_of[line_key],
                       value=round(assembly.get(total_key) or 0.0, 4))
        cell.font = Font(bold=True)
        cell.number_format = money_fmt
    for col in range(1, len(COLUMNS) + 1):
        ws.cell(row=r, column=col).fill = _TOTAL_FILL

    _autosize(ws, header_row)
    return wb


def workbook_bytes(assembly: dict, currency: str = "SEK") -> bytes:
    buffer = BytesIO()
    build_workbook(assembly, currency).save(buffer)
    return buffer.getvalue()


def _autosize(ws, header_row: int, max_width: int = 60) -> None:
    """Size columns to the header + data rows (ignores the merged title block)."""
    for col in range(1, len(COLUMNS) + 1):
        letter = get_column_letter(col)
        width = max(
            (len(str(ws.cell(row=row, column=col).value))
             for row in range(header_row, ws.max_row + 1)
             if ws.cell(row=row, column=col).value is not None),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max_width, max(10, width + 2))
